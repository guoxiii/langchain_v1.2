# src/email_workflow/nodes/generate_report.py

"""報表產生節點 — 把處理結果變成漂亮的 Excel"""

from __future__ import annotations
import json
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv

load_dotenv()

from pydantic import BaseModel, Field
from langchain.chat_models import init_chat_model
from src.email_workflow.state import WorkflowState

class DailyReportSummary(BaseModel):
    """每日報表摘要的結構化格式"""
    date: str = Field(description="報表日期")
    total_emails: int = Field(description="郵件總數")
    complaint_count: int = Field(description="客訴數量")
    quote_count: int = Field(description="報價請求數量")
    general_count: int = Field(description="一般郵件數量")
    spam_count: int = Field(description="垃圾信數量")
    high_priority_count: int = Field(description="高優先級數量")
    pending_approval_count: int = Field(description="待審批數量")
    executive_summary: str = Field(
        description="給主管看的一段話摘要（100 字以內）"
    )
    action_items: list[str] = Field(
        description="需要跟進的行動項目清單"
    )

async def generate_report(state: WorkflowState) -> dict:
    """
    工作流報表站：彙整所有處理結果，產生結構化報表
    想像一下，一整天的工作做完了，
    現在要把成果寫成一份簡潔明瞭的日報。
    LLM 幫你寫摘要，Structured Output 幫你對齊格式，
    最後 openpyxl 幫你把它變成 Excel。
    三個幫手各司其職，一氣呵成。
    """
    model = init_chat_model("anthropic:claude-sonnet-4-6")
    structured_model = model.with_structured_output(DailyReportSummary)

    # 準備郵件摘要給 LLM 分析
    email_summaries = []

    for email in state.emails:
        email_summaries.append(
            f"- [{email.category}][{email.priority}] "
            f"From: {email.sender} | Subject: {email.subject} | "
            f"Action: {email.action_taken[:100]}..."
        )

    summaries_text = "\n".join(email_summaries)

    # 讓 LLM 產生結構化的報表摘要
    report = await structured_model.ainvoke(
        f"""以下是今日處理的郵件清單，請產生一份每日報表摘要：

{summaries_text}

統計資訊：
- 總計：{state.stats['total']} 封
- 客訴：{state.stats['complaints']} 封
- 報價：{state.stats['quotes']} 封
- 一般：{state.stats['general']} 封
- 垃圾信：{state.stats['spam']} 封

請特別注意高優先級的項目，並列出需要跟進的行動項目。"""
    )

    # 產生 Excel 報表
    excel_path = _generate_excel(report, state.emails)

    return {
        "report_data": report.model_dump(),
        "workflow_status": "reviewing",
    }

def _generate_excel(
    report: DailyReportSummary,
    emails: list,
) -> str:
    """
    用 openpyxl 產生 Excel 報表
    包含兩個工作表：
    1. 摘要頁：統計數據和主管摘要
    2. 明細頁：每封郵件的處理細節
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    # === 工作表 1：摘要 ===
    ws_summary = wb.active
    ws_summary.title = "每日摘要"

    # 標題
    ws_summary["A1"] = "📧 郵件處理每日報表"
    ws_summary["A1"].font = Font(size=16, bold=True)
    ws_summary["A2"] = f"日期：{report.date}"
    ws_summary["A3"] = f"主管摘要：{report.executive_summary}"

    # 統計表格
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )

    header_font = Font(color="FFFFFF", bold=True)
    headers = ["類別", "數量"]

    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=5, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    stats_rows = [
        ("📨 郵件總數", report.total_emails),
        ("🔴 客訴", report.complaint_count),
        ("🟡 報價請求", report.quote_count),
        ("🟢 一般郵件", report.general_count),
        ("⚫ 垃圾信", report.spam_count),
        ("🚨 高優先級", report.high_priority_count),
        ("⏳ 待審批", report.pending_approval_count),
    ]

    for row_idx, (label, value) in enumerate(stats_rows, 6):
        ws_summary.cell(row=row_idx, column=1, value=label)
        ws_summary.cell(row=row_idx, column=2, value=value)

    # 行動項目
    action_row = 6 + len(stats_rows) + 1
    ws_summary.cell(
        row=action_row, column=1, value="📋 待跟進行動項目"
    ).font = Font(bold=True)

    for i, item in enumerate(report.action_items, 1):
        ws_summary.cell(
            row=action_row + i, column=1, value=f"{i}. {item}"
        )

    # 調整欄寬
    ws_summary.column_dimensions["A"].width = 40
    ws_summary.column_dimensions["B"].width = 15

    # === 工作表 2：郵件明細 ===
    ws_detail = wb.create_sheet("郵件明細")

    detail_headers = [
        "寄件人", "主旨", "分類", "優先級",
        "處理動作", "需要審批", "審批狀態"
    ]

    for col, header in enumerate(detail_headers, 1):
        cell = ws_detail.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, email in enumerate(emails, 2):
        ws_detail.cell(row=row_idx, column=1, value=email.sender)
        ws_detail.cell(row=row_idx, column=2, value=email.subject)
        ws_detail.cell(row=row_idx, column=3, value=email.category)
        ws_detail.cell(row=row_idx, column=4, value=email.priority)

        ws_detail.cell(
            row=row_idx, column=5, value=email.action_taken[:200]
        )

        ws_detail.cell(
            row=row_idx, column=6,
            value="是" if email.needs_approval else "否"
        )

        ws_detail.cell(
            row=row_idx, column=7, value=email.approval_status
        )

    # 調整明細頁欄寬
    for col_letter in ["A", "B", "C", "D", "E", "F", "G"]:
        ws_detail.column_dimensions[col_letter].width = 20

    ws_detail.column_dimensions["E"].width = 50

    # 儲存
    output_dir = Path("reports")
    output_dir.mkdir(exist_ok=True)
    filename = f"daily_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
    filepath = output_dir / filename
    wb.save(str(filepath))

    print(f"✅ 報表已產生：{filepath}")
    return str(filepath)
